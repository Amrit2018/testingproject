import openpyxl
from openpyxl.styles import PatternFill
def get_row(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_row

def get_col(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_column

def read_val(file, sheet, r, c):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.cell(r,c).value

def write_val(file, sheet, r, c, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    sheet.cell(r,c).value = data
    workbook.save(file)

def fillGreenColor(file, sheet, r, c):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    greenFill = PatternFill(start_color='rgb(255, 105, 180)', end_color='rgb(255, 105, 180)', fill_type="solid")
    sheet.cell(r, c).fill = greenFill
    workbook.save(file)
#98ff98

def fillRedColor(file, sheet, r, c):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    redFill = PatternFill(start_color='rgb(255, 0, 0)', end_color='rgb(255, 0, 0)', fill_type="solid")
    sheet.cell(r, c).fill = redFill
    workbook.save(file)
#ff0000