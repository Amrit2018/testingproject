import openpyxl

file = "D:\Python\python-selenium\seleniumPavan\sampletest.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1,1).value = 'Amrit'
sheet.cell(1,2).value = 'Mohit'
sheet.cell(1,3).value = 'Binay'
sheet.cell(1,4).value = 'Chandni'
sheet.cell(1,5).value = 'Sumant'

sheet.cell(2,1).value = '1010'
sheet.cell(2,2).value = '1011'
sheet.cell(2,3).value = '1012'
sheet.cell(2,4).value = '1013'
sheet.cell(2,5).value = '1014'

sheet.cell(3,1).value = 'Mumbai'
sheet.cell(3,2).value = 'Bangalore'
sheet.cell(3,3).value = 'Patna'
sheet.cell(3,4).value = 'Kharati'
sheet.cell(3,5).value = 'Jamshedpur'

sheet.cell(4,1).value = 'Automation'
sheet.cell(4,2).value = 'Developer'
sheet.cell(4,3).value = 'Retired HOD'
sheet.cell(4,4).value = 'Professor'
sheet.cell(4,5).value = 'Retired AGM'

workbook.save(file)

row = sheet.max_row
col = sheet.max_column
for r in range(1, row+1):
    for c in range(1, col+1):
        print(sheet.cell(r, c).value, end = '            ')
    print()

